from django.shortcuts import get_object_or_404, render
from rest_framework.views import APIView
from django.contrib.auth import get_user_model
from django.http import JsonResponse
from rest_framework.response import Response
from rest_framework import status
from admin_app.models import *
from customer.serializers import *
from .serializers import *
from django.contrib.auth.hashers import make_password
from django.contrib.auth import login
from rest_framework_simplejwt.tokens import RefreshToken
import logging
from datetime import datetime, timedelta
from django.contrib.auth import authenticate, login
from rest_framework.permissions import IsAuthenticated,AllowAny
from django.contrib.auth import authenticate
from rest_framework.parsers import JSONParser
from django.db.models import Sum
from django.http import QueryDict
import requests
from random import randint
from botocore.exceptions import ClientError
from utils.s3connector import upload_image_to_s3, upload_to_s3
from utils.light_connector import RoomUpdateView
from sleeping_pod.sleepingpod import save_pods
from django.db.models import Count
from django.db import transaction
from datetime import datetime, timedelta
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
json_file_path = 'utils/tax.json'
logger = logging.getLogger(__name__)

class BillSubmitAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, bill_number):
        get_bill = Billreport.objects.filter(bill_number=bill_number, is_deleted=False, status='Active')
        serializer = BillReportSerializer(get_bill, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
        
    def post(self, request):
        try:
            for value in request.data['service']:
                if value['service'] !=1:
                    chk_wallet = Wallet.objects.filter(user=3,is_deleted=False).first()
                    if chk_wallet:
                        balance = int(value['total'])+chk_wallet.balance
                        Wallet.objects.filter(pk=chk_wallet.id).update(amount=value['total'],balance=balance)
                        transaction_data = {
                            'user': request.user.id,
                            'amount': value['total'],
                            'balance': balance ,
                            'requested_by':request.user.id,
                            'transaction_type':1
                            }
                        transactionSerializer = WalletTransactionSerializer(data=transaction_data)
                        if transactionSerializer.is_valid():
                            transactionSerializer.save()
                        else:
                            return Response(transactionSerializer.errors, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        data = {
                        'user': request.user.id,
                        'amount': value['total'],
                        'balance': value['total'] ,
                        'requested_by':request.user.id
                        }
                        serializer = WalletSerializer(data=data)
                        if serializer.is_valid():
                            serializer.save()
                            transaction_data = {
                                'user': request.user.id,
                                'amount': value['total'],
                                'balance': value['total'] ,
                                'requested_by':request.user.id,
                                'transaction_type':1
                            }
                            transactionSerializer = WalletTransactionSerializer(data=transaction_data)
                            if transactionSerializer.is_valid():
                                transactionSerializer.save()
                            else:
                                return Response(transactionSerializer.errors, status=status.HTTP_400_BAD_REQUEST)
                        else:
                                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


                if value['room_numbers']:
                    room_update_data = value['room_numbers']
                    room_update_view = RoomUpdateView()
                    room_update_response = room_update_view.post(room_update_data)
                
                # save bill report
                report_data = {
                    'user': request.user.id,
                    'bill_number':request.data['bill_number'],
                    'service': value['service'],
                    'serviceType': value['serviceType'],
                    'quantity':value['number'],
                    'rate':value['rate'],
                    'total':value['total'],
                    'payment_mode':request.data['payment_mode'],
                    'split': request.data['split'],
                    'room_numbers': value['room_numbers'],
                    'sleepingpod_numbers': value['sleepingpod_numbers'],
                    'hours': value['hours'],

                }
                reportSerializer = BillReportSerializer(data=report_data)
                if reportSerializer.is_valid():
                    reportSerializer.save()
                    if value['service'] ==1:
                        checkin_time = request.data['checkin_time']
                        checkin_time_obj = datetime.strptime(str(checkin_time), '%H:%M:%S').time()
                        checkout_datetime = datetime.combine(datetime.today(), checkin_time_obj) + timedelta(hours=4)
                        checkout_time = checkout_datetime.strftime('%H:%M:%S')
                        checkout_time_obj = datetime.strptime(str(checkout_time), '%H:%M:%S').time()
                        event_data = {
                            'user': request.user.id,
                            'service': value['service'],
                            'serviceType': value['serviceType'],
                            'number':value['number'],
                            'sleepingpod_numbers': value['sleepingpod_numbers'],
                            'hours': value['hours'],
                            'source':'b',
                            'checkin_time':request.data['checkin_time'],
                            'checkout_time':checkout_time_obj,
                            'date':timezone.now().date(),
                            'sleepingpod_package_id':value['sleepingpod_package_id']
                        }
                        events_save=save_pods(event_data,request.data['mobile'])
                        if events_save == 400:
                            return Response({'message': 'bill report cant created!'}, status=status.HTTP_400_BAD_REQUEST)


                else:
                    return Response({'message': 'bill report cant created!'}, status=status.HTTP_400_BAD_REQUEST)
            return Response({'message': 'Bill submited successfully!'}, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            logger.error(f"Error occurred in BillSubmitAPIView POST: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class EventReportAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            service_type_id = request.GET.get('service_type')
            if service_type_id:
                events = Event.objects.filter(is_deleted=False, serviceType=service_type_id)
            else:
                events = Event.objects.filter(is_deleted=False)
            serializer = EventReportSerializer(events, many=True)
            return Response(serializer.data,status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error occurred in EventReportAPIView GET: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class EventReportExportAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            service_type_id = request.GET.get('service_type')
            if service_type_id:
                events = Event.objects.filter(is_deleted=False, serviceType=service_type_id)
            else:
                events = Event.objects.filter(is_deleted=False)
            serializer = EventReportSerializer(events, many=True)
            # Create an Excel workbook and sheet
            data = serializer.data
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Event Report"

            # Write headers
            headers = ["created_on", "user", "checkin_time", "checkout_time", "service", "serviceType", "number"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}1"] = header

            # Write data
            for row_num, item in enumerate(data, 2):
                sheet[f"A{row_num}"] = item["created_on"]
                details = item["user"]
                details = item["checkin_time"]
                details = item["checkout_time"]
                details = item["service"]
                details = item["serviceType"]
                details = item["number"]
                detail_str = "\n".join([str(detail) for detail in details])
                sheet[f"B{row_num}"] = detail_str

            # Save workbook to a BytesIO buffer
            from io import BytesIO
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            # Create the HTTP response with the Excel file
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Bill_Report.xlsx'

            return response


            # return Response(serializer.data,status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error occurred in EventReportAPIView GET: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
class UserReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            type = request.GET.get('type')
            users = User.objects.exclude(role=1)
            if type:
                if type == 'subscription':
                    users = users.filter(subscription__is_deleted=False).exclude(role=1)
                elif type == 'wallet':
                    users = users.filter(wallet__is_deleted=False).exclude(role=1)
            serializer = UserReportSerializer(users, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error occurred in UserReportAPIView GET: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
class WalletAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            chk_wallet = Wallet.objects.filter(user=request.data['user'],is_deleted=False).first()
            if chk_wallet:
                balance = int(request.data['amount'])+chk_wallet.balance
                Wallet.objects.filter(pk=chk_wallet.id).update(amount=request.data['amount'],balance=balance)
                transaction_data = {
                    'user': request.data['user'],
                    'amount': request.data['amount'],
                    'balance': balance ,
                    'requested_by':request.user.id,
                    'transaction_type':1
                    }
                transactionSerializer = WalletTransactionSerializer(data=transaction_data)
                if transactionSerializer.is_valid():
                    transactionSerializer.save()
                    return Response({'message': 'Wallet added successfully!'}, status=status.HTTP_201_CREATED)
                    
                else:
                    return Response(transactionSerializer.errors, status=status.HTTP_400_BAD_REQUEST)
            else:
                data = {
                    'user': request.data['user'],
                    'amount': request.data['amount'],
                    'balance': request.data['amount'] ,
                    'requested_by':request.user.id
                    }
                serializer = WalletSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    transaction_data = {
                    'user': request.data['user'],
                    'amount': request.data['amount'],
                    'balance': request.data['amount'] ,
                    'requested_by':request.user.id,
                    'transaction_type':1
                    }
                    transactionSerializer = WalletTransactionSerializer(data=transaction_data)
                    if transactionSerializer.is_valid():
                        transactionSerializer.save()
                        return Response({'message': 'Wallet added successfully!'}, status=status.HTTP_201_CREATED)
                    else:
                        return Response(transactionSerializer.errors, status=status.HTTP_400_BAD_REQUEST)
                else:    
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            logger.error(f"Error occurred in WalletAPIView POST: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class SignupAPIView(APIView):
    def post(self, request):
        try:
            role = int(request.data.get('role'))
            role_mapping = dict(SELECTROLE)
            if role not in role_mapping:
                return Response({"error": "Invalid role specified"}, status=status.HTTP_400_BAD_REQUEST)
            
            data = {
                'username': request.data['username'],
                'email': request.data['email'],
                'password': request.data['password'],
                'role': role,
            }

            user = User.objects.create_user(**data)
            user.is_active = True
            user.save()
            
            role_name = role_mapping[role].upper()
            user.uid = f"TRV-{role_name}-{user.id}"
            user.save()

            return Response({"message": "User created successfully"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            logger.error(f"Error occurred in SignupAPIView POST: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
      
class ForgotpasswordAPIView(APIView):
    def post(self,request):
        try:
            email = request.data.get('email')
            password = request.data.get('password')
            confirm_password = request.data.get('confirm_password')
            
            if not email:
                return Response({"error":"Email is required"},status=status.HTTP_400_BAD_REQUEST)
            
            if password != confirm_password:
                return Response({"error":"Password mismatch"},status=status.HTTP_400_BAD_REQUEST)
            
            user = User.objects.get(email=email)
            user.set_password(password)
            user.save()
            
            return Response({"message":"Password Reset success"},status=status.HTTP_201_CREATED)
        except User.DoesNotExist:
            return Response({"error":"User with this email does not exist"},status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error ocurred in ForgetpasswordAPIView POST: {e}")
            return Response({"error":"An error ocurred"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        
class Signin(serializers.Serializer):
    username = serializers.CharField()
    password = serializers.CharField()
class SigninAPIView(APIView):
    permission_classes = [AllowAny]
    def post(self,request): 
        try:
            username = request.data.get('username')
            password = request.data.get('password')
            user = authenticate(request,username=username,password=password)
            if user is not None:
                refresh = RefreshToken.for_user(user)
                response ={
                    "access_token":str(refresh.access_token),
                    "refresh_token":str(refresh),
                    "status_code":status.HTTP_200_OK
                }
                return Response(response,status=status.HTTP_200_OK)
            else:
                return Response({"message":"Incorrect inputs"},status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error occured in signin POST:{e}")
            return Response({"error":"An error occured"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class GroupedSerializer(serializers.Serializer):
    bill_number = serializers.CharField()
    details = serializers.ListField(child=serializers.DictField())

    class Meta:
        fields = ['bill_number', 'details']

class BillReportAPIView(APIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            
            service_type_id = request.GET.get('service_type')
            if service_type_id:
                report = Billreport.objects.filter(is_deleted=False, serviceType=service_type_id)
            else:
                report = Billreport.objects.filter(is_deleted=False)

            # Group by bill_number and serialize each group
            grouped_data = {}
            for item in report:
                if item.bill_number not in grouped_data:
                    grouped_data[item.bill_number] = []
                serializer = BillReportGetSerializer(item)
                grouped_data[item.bill_number].append(serializer.data)

            # Create a list of dictionaries for the GroupedSerializer
            grouped_list = [{'bill_number': key, 'details': value} for key, value in grouped_data.items()]

            serializer = GroupedSerializer(data=grouped_list, many=True)
            serializer.is_valid(raise_exception=True)

            return Response(serializer.validated_data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error occurred in BillReportAPIView GET: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR) 
        
class BillReportExportAPIView(APIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            fromdate = request.GET.get('fromdate')
            todate = request.GET.get('todate')

            if fromdate and todate:
                report = Billreport.objects.filter(is_deleted=False, created_on__gte=fromdate, created_on__lte=todate)
            else:
                report = Billreport.objects.filter(is_deleted=False)

            # Group by bill_number and serialize each group
            grouped_data = {}
            for item in report:
                if item.bill_number not in grouped_data:
                    grouped_data[item.bill_number] = []
                serializer = BillReportGetSerializer(item)
                grouped_data[item.bill_number].append(serializer.data)

            # Create a list of dictionaries for the GroupedSerializer
            grouped_list = [{'bill_number': key, 'details': value} for key, value in grouped_data.items()]

            # Serialize grouped data
            serializer = GroupedSerializer(data=grouped_list, many=True)
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data

            # Create an Excel workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Bill Report"

            # Write headers
            headers = ["Bill Number", "Details"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}1"] = header

            # Write data
            for row_num, item in enumerate(data, 2):
                sheet[f"A{row_num}"] = item["bill_number"]
                details = item["details"]
                detail_str = "\n".join([str(detail) for detail in details])
                sheet[f"B{row_num}"] = detail_str

            # Save workbook to a BytesIO buffer
            from io import BytesIO
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            # Create the HTTP response with the Excel file
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Bill_Report.xlsx'

            return response

        except Exception as e:
            logger.error(f"Error occurred in BillReportExportAPIView GET: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class RefundAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            if request.user.role == 1:
                bill_number = request.data['bill_number']
                amount = request.data['amount']
                get_user = Billreport.objects.filter(bill_number=bill_number,is_deleted=False).first()
                bill_return = Billreport.objects.filter(bill_number=bill_number,is_deleted=False).update(status='Return')
                if bill_return:
                    chk_wallet = Wallet.objects.filter(user=get_user.user_id,is_deleted=False,requested_by=3).first()
                    if chk_wallet:
                        if chk_wallet.balance < int(amount):
                            return Response({"message": "The biller's wallet does not contain the correct refund amount!"}, status=status.HTTP_400_BAD_REQUEST)
                        balance = chk_wallet.balance-int(amount)
                        Wallet.objects.filter(pk=chk_wallet.id).update(amount=amount,balance=balance)
                        transaction_data = {
                            'user': get_user.user_id,
                            'amount': amount,
                            'balance': balance ,
                            'requested_by':request.user.id,
                            'transaction_type':0
                        }
                        transactionSerializer = WalletTransactionSerializer(data=transaction_data)
                        if transactionSerializer.is_valid():
                            transactionSerializer.save()
                            return Response({'message': 'Bill returned successfully!'}, status=status.HTTP_201_CREATED)
                    else:
                        return Response(transactionSerializer.errors, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'message': 'Permissin denied, please contact with super admin!'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error occurred in RefundAPIView POST: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)  
        
class BillUpdateAPIView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            if request.user.role == 1:
                get_user = Billreport.objects.filter(bill_number=request.data['bill_number'],is_deleted=False).first()
                old_total = Billreport.objects.filter(bill_number=request.data['bill_number'],is_deleted=False).aggregate(sum_amount=Sum('total'))
                chk_oldwallet = Wallet.objects.filter(user=get_user.user_id,is_deleted=False).first()
                if chk_oldwallet:
                    if chk_oldwallet.balance < old_total['sum_amount']:
                        return Response({"message": "The biller's wallet does not contain the correct refund amount!!"}, status=status.HTTP_400_BAD_REQUEST)
                    old_balance = chk_oldwallet.balance-old_total['sum_amount']
                    old_wallet = Wallet.objects.filter(pk=chk_oldwallet.id).update(amount=old_total['sum_amount'],balance=old_balance)    
                    if old_wallet == 1:
                        oldtransaction_data = {
                        'user': get_user.user_id,
                        'amount': old_total['sum_amount'],
                        'balance': old_balance ,
                        'requested_by':request.user.id,
                        'transaction_type':0
                        }
                    oldtransactionSerializer = WalletTransactionSerializer(data=oldtransaction_data)
                    if oldtransactionSerializer.is_valid():
                        oldtransactionSerializer.save()
                        chk_wallet = Wallet.objects.filter(user=get_user.user_id,is_deleted=False).first()
                        balance = int(request.data['amount'])+chk_wallet.balance
                        Wallet.objects.filter(pk=chk_wallet.id).update(amount=request.data['amount'],balance=balance)
                        transaction_data = {
                            'user': get_user.user_id,
                            'amount': request.data['amount'],
                            'balance': balance ,
                            'requested_by':request.user.id,
                            'transaction_type':1
                            }
                        transactionSerializer = WalletTransactionSerializer(data=transaction_data)
                        if transactionSerializer.is_valid():
                            transactionSerializer.save()
                            Billreport.objects.filter(bill_number=request.data['bill_number'],is_deleted=False).update(is_deleted=True)
                            for value in request.data['service']:
                                room_update_data = value['room_numbers']
                                room_update_view = RoomUpdateView()
                                room_update_response = room_update_view.post(room_update_data)
                                
                                # save bill report
                                report_data = {
                                    'user': get_user.user_id,
                                    'bill_number':request.data['bill_number'],
                                    'service': value['service'],
                                    'serviceType': value['serviceType'],
                                    'quantity':value['number'],
                                    'rate':value['rate'],
                                    'total':value['total'],
                                    'payment_mode':request.data['payment_mode'],
                                    'split': request.data['split'],
                                    'room_numbers': value['room_numbers']
                                }
                                reportSerializer = BillReportSerializer(data=report_data)
                                if reportSerializer.is_valid():
                                    reportSerializer.save()
                                else:
                                    return Response({'message': 'bill report cant created!'}, status=status.HTTP_400_BAD_REQUEST)
                            return Response({'message': 'Bill updated successfully!'}, status=status.HTTP_201_CREATED)
                        else:
                                return Response(transactionSerializer.errors, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        return Response(oldtransactionSerializer.errors, status=status.HTTP_400_BAD_REQUEST)  
            else:
                return Response({'message': 'Permissin denied, please contact with super admin!'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error occurred in BillSubmitAPIView POST: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)       


class UserCheckApiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            mobile_number = request.GET.get('mobile_number')
            user = User.objects.get(mobile_number=mobile_number,role = 3)
            if user:
                serializer = UserProfileSerializer(instance=user)
                data = serializer.data
                
                print("data==== = = =", data)
                proof_data = data['idproof']
                if proof_data:
                    proof_urls = ["https://travlounge-dev-bucket.s3.amazonaws.com" + proof['id_proof'] for proof in proof_data]
                    data['idproof'] = proof_urls
                else:
                    data['idproof'] = []


                return Response(data, status=status.HTTP_200_OK)
            else:
                return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error occurred in UserCheckApiView GET: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        users_data = request.data
        print("request = = = = ",users_data)
        responses = {}
        
        for user_data in users_data:
            name = user_data.get('name')
            mobile_number = user_data.get('mobile_number')
            gender = user_data.get('gender')

            user = User.objects.filter(mobile_number=mobile_number, role=3).first()
            if user:
                responses[mobile_number] = {"message": "User with mobile number already exists"}
                continue

            try:
                last_record = User.objects.latest('id')
                lastid = last_record.id + 1
            except User.DoesNotExist:
                lastid = 1

            uid = "TRV-CUSTOMER-" + str(lastid)

            create = User.objects.create_user(
                name=name,
                username=str(mobile_number),
                mobile_number=str(mobile_number),
                password=str(mobile_number),
                gender=gender,
                role=3,
                uid=uid
            )

            if create:
                try:
                    user_id = create.id
                    # Assuming only one image per user
                    image = request.FILES.get(f'idProof{user_id}')
                    if image:
                        upload = upload_to_s3(image, user_id)
                        if upload.get("error"):
                            responses[mobile_number] = {"message": "User created incompletely", "error": upload.get("error")}
                            continue
                    else:
                        responses[mobile_number] = {"message": "User created incompletely"}
                except Exception as e:
                    print(e)
                    responses[mobile_number] = {"message": "Error occurred while processing user"}
                    continue

        return Response(responses, status=status.HTTP_201_CREATED)


    def put(self, request, pk):
        try:
            with transaction.atomic():
                # Delete all entries for the specified user in the idproof table
                IdProof.objects.filter(user_id=pk, is_deleted = False).update(is_deleted = True)

                user = User.objects.filter(id=pk).first()
                if not user:
                    return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

                images = request.FILES.getlist('idProof')
                
                if images:
                    for image in images:
                        upload = upload_to_s3(image, user.id)
                else:
                    upload = {"error":"no image file uploaded"}
                    
                if upload.get("error"):
                    return Response({"message": "document upload unsuccessfull", "error":upload["error"]}, status=status.HTTP_409_CONFLICT)

                return Response({'message': 'Document uploaded successfully'})
        except Exception as e:
            logger.error(f"Error occurred in UserCheckApiView PUT: {e}")
            return Response({"error": "An error occurred"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
        
